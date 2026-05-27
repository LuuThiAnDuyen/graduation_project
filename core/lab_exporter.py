"""
core/lab_exporter.py
---------------------
Export kết quả Prompt Lab ra file Excel.

Sheets:
  1. 📊 Dimension Scores  — điểm 6 chiều từng variant
  2. 📋 All Test Cases    — toàn bộ TC của các variant
  3. 🔍 Analysis Notes    — strengths, weaknesses, issues
  4. 🧠 LLM Judge         — semantic scores (chỉ khi used_llm_judge=True)
  5. ⚠️ Error Report      — báo cáo lỗi 7 loại (chỉ khi có error_report)

Đã xoá: 🏆 Leaderboard, 🏁 Win Matrix
"""

from __future__ import annotations
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style helpers ─────────────────────────────────────────────────────────────


def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr_style():
    return (
        PatternFill("solid", fgColor="1F4E79"),
        Font(bold=True, color="FFFFFF", size=11),
        Alignment(vertical="top", wrap_text=True, horizontal="left"),
    )


def _body_align():
    return Alignment(vertical="top", wrap_text=True, horizontal="left")


_GRADE_COLORS = {
    "A": "C6EFCE",
    "B": "EBFACC",
    "C": "FFEB9C",
    "D": "FFCC7A",
    "F": "FFC7CE",
}
_SCORE_FILLS = {
    (85, 101): PatternFill("solid", fgColor="C6EFCE"),
    (70, 85): PatternFill("solid", fgColor="EBFACC"),
    (55, 70): PatternFill("solid", fgColor="FFEB9C"),
    (40, 55): PatternFill("solid", fgColor="FFCC7A"),
    (0, 40): PatternFill("solid", fgColor="FFC7CE"),
}
_SEVERITY_FILLS = {
    "Critical": PatternFill("solid", fgColor="FFC7CE"),
    "Warning": PatternFill("solid", fgColor="FFEB9C"),
    "Info": PatternFill("solid", fgColor="DDEBF7"),
}


def _score_fill(score: float) -> PatternFill:
    for (lo, hi), fill in _SCORE_FILLS.items():
        if lo <= score < hi:
            return fill
    return PatternFill("solid", fgColor="F0F0F0")


def _write_headers(ws, headers: list[str], widths: list[int]) -> None:
    hdr_fill, hdr_font, hdr_align = _hdr_style()
    border = _thin()
    ws.append(headers)
    for c, w in enumerate(widths, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = (
            hdr_fill,
            hdr_font,
            hdr_align,
            border,
        )
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


# ── Sheet 1: Dimension Scores ─────────────────────────────────────────────────


def _build_dimensions(ws, lab_result) -> None:
    from core.tc_evaluator import DIMENSION_LABELS

    dim_names = list(DIMENSION_LABELS.values())
    winner_map = lab_result.dimension_winner_map or {}

    headers = (
        ["Variant"]
        + [f"{d} {'🏆' if winner_map.get(d) else ''}" for d in dim_names]
        + ["Overall Score", "Grade"]
    )
    widths = [22] + [18] * len(dim_names) + [14, 8]
    _write_headers(ws, headers, widths)
    border = _thin()

    for run in lab_result.leaderboard:
        ev = run.evaluation
        dim_map = {d.name: d.percentage for d in ev.dimensions}
        row = (
            [run.variant_name]
            + [dim_map.get(d, 0) for d in dim_names]
            + [round(ev.overall_score, 1), ev.grade]
        )
        ws.append(row)
        r = ws.max_row
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(val, (int, float)) and 2 <= c <= len(dim_names) + 1:
                cell.fill = _score_fill(float(val))
                dim_idx = c - 2
                if dim_idx < len(dim_names):
                    bold = winner_map.get(dim_names[dim_idx]) == run.variant_name
                    cell.font = Font(
                        size=10, bold=bold, color="1F4E79" if bold else "000000"
                    )
                else:
                    cell.font = Font(size=10)
            else:
                cell.font = Font(size=10)
            cell.alignment = _body_align()
            cell.border = border
    ws.auto_filter.ref = ws.dimensions


# ── Sheet 2: All Test Cases ───────────────────────────────────────────────────


def _build_all_tc(ws, lab_result) -> None:
    headers = [
        "Variant",
        "TC ID",
        "Title",
        "Type",
        "Priority",
        "Precondition",
        "Steps",
        "Expected Result",
        "DB Query",
        "Test Data Ref",
    ]
    widths = [18, 8, 35, 8, 10, 30, 45, 40, 30, 12]
    _write_headers(ws, headers, widths)
    border = _thin()

    PRIORITY_FILLS = {
        "High": PatternFill("solid", fgColor="FFE0E0"),
        "Medium": PatternFill("solid", fgColor="FFF9E0"),
        "Low": PatternFill("solid", fgColor="E8F5E9"),
    }

    for run in lab_result.leaderboard:
        for tc in run.evaluation.raw_tc_list:
            row = [
                run.variant_name,
                tc.get("id", ""),
                tc.get("title", ""),
                tc.get("coverage_type", ""),
                tc.get("priority", ""),
                tc.get("precondition", ""),
                tc.get("steps_text", ""),
                tc.get("expected_result", ""),
                tc.get("db_query", ""),
                tc.get("test_data_ref", ""),
            ]
            ws.append(row)
            pfill = PRIORITY_FILLS.get(
                tc.get("priority", "Medium"), PRIORITY_FILLS["Medium"]
            )
            r = ws.max_row
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = pfill
                cell.font = Font(size=10)
                cell.alignment = _body_align()
                cell.border = border
    ws.auto_filter.ref = ws.dimensions


# ── Sheet 3: Analysis Notes ───────────────────────────────────────────────────


def _build_analysis(ws, lab_result) -> None:
    headers = [
        "Rank",
        "Variant",
        "Description",
        "Tags",
        "Strengths",
        "Weaknesses",
        "Issues",
        "Verdict",
    ]
    widths = [6, 22, 40, 30, 40, 40, 50, 50]
    _write_headers(ws, headers, widths)
    border = _thin()

    for rank, run in enumerate(lab_result.leaderboard, 1):
        ev = run.evaluation
        all_issues = []
        for d in ev.dimensions:
            all_issues.extend(d.issues)
        row = [
            rank,
            run.variant_name,
            run.variant_description,
            ", ".join(run.variant_tags),
            "\n".join(f"• {s}" for s in ev.strengths),
            "\n".join(f"• {w}" for w in ev.weaknesses),
            "\n".join(f"• {i}" for i in all_issues[:10]),
            ev.recommendation,
        ]
        ws.append(row)
        r = ws.max_row
        row_fill = PatternFill("solid", fgColor=_GRADE_COLORS.get(ev.grade, "F0F0F0"))
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = row_fill
            cell.font = Font(size=10)
            cell.alignment = _body_align()
            cell.border = border
        ws.row_dimensions[r].height = max(
            60, 15 * max(len(ev.strengths), len(ev.weaknesses), 3)
        )
    ws.auto_filter.ref = ws.dimensions


# ── Sheet 4: LLM Judge ────────────────────────────────────────────────────────


def _build_llm_judge(ws, lab_result) -> None:
    headers = [
        "Variant",
        "Semantic Coverage",
        "Step Clarity",
        "Expected Result Quality",
        "Test Data Realism",
        "Negative Case Quality",
        "Judge Composite",
        "Verdict",
    ]
    widths = [22, 18, 14, 22, 18, 20, 16, 60]
    _write_headers(ws, headers, widths)
    border = _thin()

    for run in lab_result.leaderboard:
        ev = run.evaluation
        js = ev.llm_judge_score
        if js is None or not js.judge_available:
            row = [run.variant_name] + ["N/A"] * 6 + ["Judge not available"]
        else:
            row = [
                run.variant_name,
                js.semantic_coverage,
                js.step_clarity,
                js.expected_result_quality,
                js.test_data_realism,
                js.negative_case_quality,
                round(js.composite_score, 1),
                js.verdict,
            ]
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            val = row[c - 1]
            if isinstance(val, (int, float)) and 2 <= c <= 7:
                cell.fill = _score_fill(float(val))
            else:
                cell.fill = PatternFill("solid", fgColor="F8F8F8")
            cell.font = Font(size=10)
            cell.alignment = _body_align()
            cell.border = border
    ws.auto_filter.ref = ws.dimensions


# ── Sheet 5: Error Report ─────────────────────────────────────────────────────


def _build_error_report(ws, lab_result) -> None:
    """
    Sheet tổng hợp lỗi của tất cả variants đã chạy.
    Columns: Variant | Error Type | TC ID | Severity | Source | Description | Suggestion
    """
    from core.error_analyzer import ERROR_TYPES

    headers = [
        "Variant",
        "Error Type",
        "Mô tả lỗi",
        "TC ID",
        "Severity",
        "Nguồn",
        "Mô tả chi tiết",
        "Gợi ý sửa",
    ]
    widths = [20, 10, 38, 12, 12, 8, 55, 50]
    _write_headers(ws, headers, widths)
    border = _thin()

    has_any = False
    for run in lab_result.leaderboard:
        ev = run.evaluation
        if ev.error_report is None:
            continue
        for error in ev.error_report.all_errors:
            row = [
                run.variant_name,
                error.error_type,
                ERROR_TYPES.get(error.error_type, ""),
                error.tc_id,
                error.severity,
                "LLM" if error.source == "llm" else "Rule",
                error.description,
                error.suggestion,
            ]
            ws.append(row)
            r = ws.max_row
            sev_fill = _SEVERITY_FILLS.get(
                error.severity, PatternFill("solid", fgColor="F8F8F8")
            )
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = sev_fill
                cell.font = Font(size=10)
                cell.alignment = _body_align()
                cell.border = border
            ws.row_dimensions[r].height = 30
            has_any = True

    if not has_any:
        ws.append(["Không có dữ liệu Error Analysis.", "", "", "", "", "", "", ""])
        ws.append(
            [
                "Bật toggle 🔍 Error Analysis trước khi chạy Lab.",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )

    ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.tabColor = "FF9800"


# ── Public API ────────────────────────────────────────────────────────────────


def lab_to_excel(lab_result) -> bytes:
    """
    Export LabResult → Excel bytes.

    Sheets luôn có:
      1. 📊 Dimension Scores
      2. 📋 All Test Cases
      3. 🔍 Analysis Notes

    Sheet thêm nếu used_llm_judge=True:
      4. 🧠 LLM Judge

    Sheet thêm nếu có error_report:
      5. ⚠️ Error Report
    """
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "📊 Dimension Scores"
    _build_dimensions(ws1, lab_result)

    ws2 = wb.create_sheet("📋 All Test Cases")
    _build_all_tc(ws2, lab_result)

    ws3 = wb.create_sheet("🔍 Analysis Notes")
    _build_analysis(ws3, lab_result)

    if lab_result.used_llm_judge:
        ws4 = wb.create_sheet("🧠 LLM Judge")
        _build_llm_judge(ws4, lab_result)

    # Error Report: thêm nếu ít nhất 1 variant có error_report
    has_error_report = any(
        r.evaluation and r.evaluation.error_report is not None
        for r in lab_result.successful_runs
    )
    if has_error_report:
        ws5 = wb.create_sheet("⚠️ Error Report")
        _build_error_report(ws5, lab_result)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
