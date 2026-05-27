"""
core/excel_exporter.py
-----------------------
Export test cases + test data sang Excel.

Hai chế độ:
  • to_excel(test_cases, test_data_set)
      – Sheet per feature           : mỗi chức năng 1 sheet riêng
      – Sheet "Data – <feat>"       : test data đi kèm ngay sau sheet TC
      – Sheet "📋 Quick Summary"    : bảng ngắn gọn Sheet|Feature|TC|Ghi chú  ← MỚI
      – Sheet "📊 Coverage Summary" : bảng tổng hợp chi tiết (cuối cùng)

  • to_excel_single(test_cases, test_data_set, sheet_name)
      – 3 sheet: TC + TD + Summary (dùng nội bộ / fallback)

Thay đổi:
  • Thêm sheet "📋 Quick Summary" (ngay trước Coverage Summary) — bảng ngắn:
    Sheet name | Feature name | Số TC | Ghi chú (coverage % + highlights)
  • Element & Locator: nếu một element name giống dòng liền trên (trong cùng TC)
    thì thay bằng chuỗi rỗng để tránh lặp lại.
  • _normalise_feature_group(): gom các feature_group bị LLM split sai về đúng module.
"""

from __future__ import annotations
import re
from io import BytesIO
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style helpers ────────────────────────────────────────────────────────────


def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr_style():
    return (
        PatternFill("solid", fgColor="1F4E79"),
        Font(bold=True, color="FFFFFF", size=11, name="Arial"),
        Alignment(vertical="top", wrap_text=True, horizontal="left"),
    )


_PRIORITY_FILLS = {
    "High": PatternFill("solid", fgColor="FFE0E0"),
    "Medium": PatternFill("solid", fgColor="FFF9E0"),
    "Low": PatternFill("solid", fgColor="E8F5E9"),
}
_BODY_ALIGN = Alignment(vertical="top", wrap_text=True, horizontal="left")
_BODY_FONT = Font(size=10, name="Arial")
_STATUS_FILL = PatternFill("solid", fgColor="F0F0F0")
_TD_FILL = PatternFill("solid", fgColor="EAF4FB")
_ELEMENT_FILL = PatternFill("solid", fgColor="F3E8FF")

_FEATURE_HDR_COLORS = [
    "1F4E79",
    "155724",
    "5C2D91",
    "7B3F00",
    "006064",
    "4A148C",
    "1A237E",
    "BF360C",
]

# ── Outcome/type suffixes that LLMs mistakenly append to feature_group ───────
_OUTCOME_SUFFIXES = re.compile(
    r"\s*[-–—|:]\s*(?:success|fail(?:ure)?|error|valid(?:ation)?|invalid|"
    r"positive|negative|happy[\s_]?path|sad[\s_]?path|edge|boundary|"
    r"security|thành\s+công|thất\s+bại|hợp\s+lệ|không\s+hợp\s+lệ|"
    r"lỗi|kiểm\s+tra|xác\s+thực)\s*$",
    re.IGNORECASE,
)

_OUTCOME_WORDS_STANDALONE = re.compile(
    r"^(.*?)\s+(?:success(?:ful)?|fail(?:ure|ed)?|error|"
    r"valid(?:ation)?|invalid|positive|negative|happy|sad|"
    r"thành\s+công|thất\s+bại|hợp\s+lệ|không\s+hợp\s+lệ|lỗi)\s*$",
    re.IGNORECASE,
)


def _normalise_feature_group(name: str) -> str:
    name = name.strip()
    if not name:
        return "General"
    cleaned = _OUTCOME_SUFFIXES.sub("", name).strip()
    m = _OUTCOME_WORDS_STANDALONE.match(cleaned)
    if m:
        candidate = m.group(1).strip()
        if len(candidate) >= 3:
            cleaned = candidate
    return cleaned if cleaned else name


def _group_by_normalised_feature(test_cases: list[dict]) -> dict[str, list[dict]]:
    normalised: list[tuple[str, dict]] = []
    for tc in test_cases:
        raw_fg = (tc.get("feature_group") or "").strip() or "General"
        norm_fg = _normalise_feature_group(raw_fg)
        normalised.append((norm_fg, tc))

    unique_names: list[str] = list(dict.fromkeys(n for n, _ in normalised))

    merge_map: dict[str, str] = {}
    for name in unique_names:
        merge_map[name] = name

    for i, name_a in enumerate(unique_names):
        for name_b in unique_names[i + 1 :]:
            shorter, longer = (
                (name_a, name_b) if len(name_a) <= len(name_b) else (name_b, name_a)
            )
            if longer.lower().startswith(shorter.lower()):
                rest = longer[len(shorter) :]
                if not rest or rest[0] in (" ", "-", "–", "_", "/"):
                    current = merge_map[longer]
                    if len(shorter) < len(current):
                        merge_map[longer] = shorter

    groups: dict[str, list[dict]] = defaultdict(list)
    for norm_fg, tc in normalised:
        canonical = merge_map.get(norm_fg, norm_fg)
        groups[canonical].append(tc)

    return dict(groups)


# ── Column definitions ───────────────────────────────────────────────────────

_TC_COLS = [
    ("id", "ID", 10),
    ("title", "Title", 35),
    ("coverage_type", "Type", 14),
    ("priority", "Priority", 12),
    ("precondition", "Precondition", 32),
    ("steps_text", "Steps", 45),
    ("element_locator", "Element & Locator", 25),
    ("expected_result", "Expected Result", 42),
    ("actual_result", "Actual Result", 30),
    ("status_result", "Status", 12),
    ("db_query", "DB Query", 35),
    ("db_expected", "DB Expected", 25),
    ("test_data_ref", "Test Data Ref", 14),
]
_EMPTY_COLS = {9, 10}
_ELEMENT_COL = 7

_TD_COLS = [
    ("id", "TD ID", 12),
    ("description", "Description", 40),
    ("data_text", "Data (key: value)", 60),
]


# ── Element deduplication ────────────────────────────────────────────────────


def _dedup_element_locator(raw_locator: str) -> str:
    lines = raw_locator.split("\n") if raw_locator else []
    result = []
    prev = None
    for line in lines:
        stripped = line.strip()
        if prev is not None and stripped.lower() == prev.lower():
            result.append("")
        else:
            result.append(stripped)
        prev = stripped
    return "\n".join(result)


# ── Low-level helpers ────────────────────────────────────────────────────────


def _write_headers(
    ws, headers: list[str], widths: list[int], color: str = "1F4E79"
) -> None:
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    align = Alignment(vertical="top", wrap_text=True, horizontal="left")
    border = _thin()
    ws.append(headers)
    for c, w in enumerate(widths, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = fill, font, align, border
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def _write_tc_rows(ws, test_cases: list[dict]) -> None:
    keys = [c[0] for c in _TC_COLS]
    border = _thin()
    for row_idx, tc in enumerate(test_cases, start=ws.max_row + 1):
        tc_row = dict(tc)
        tc_row["element_locator"] = _dedup_element_locator(
            tc_row.get("element_locator", "")
        )
        ws.append([tc_row.get(k, "") for k in keys])
        pfill = _PRIORITY_FILLS.get(
            tc_row.get("priority", "Medium"), _PRIORITY_FILLS["Medium"]
        )
        for col_idx in range(1, len(_TC_COLS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx in _EMPTY_COLS:
                cell.fill = _STATUS_FILL
            elif col_idx == _ELEMENT_COL:
                cell.fill = _ELEMENT_FILL
            else:
                cell.fill = pfill
            cell.font = _BODY_FONT
            cell.alignment = _BODY_ALIGN
            cell.border = border


def _build_tc_sheet(ws, test_cases: list[dict], header_color: str = "1F4E79") -> None:
    headers = [c[1] for c in _TC_COLS]
    widths = [c[2] for c in _TC_COLS]
    _write_headers(ws, headers, widths, color=header_color)
    _write_tc_rows(ws, test_cases)
    ws.auto_filter.ref = ws.dimensions


def _build_td_sheet(ws, test_data_set: list[dict]) -> None:
    headers = [c[1] for c in _TD_COLS]
    widths = [c[2] for c in _TD_COLS]
    keys = [c[0] for c in _TD_COLS]
    border = _thin()
    _write_headers(ws, headers, widths)
    for row_idx, td in enumerate(test_data_set, start=2):
        ws.append([td.get(k, "") for k in keys])
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill, cell.font, cell.alignment, cell.border = (
                _TD_FILL,
                _BODY_FONT,
                _BODY_ALIGN,
                border,
            )
    ws.auto_filter.ref = ws.dimensions


# ── Feature grouping ─────────────────────────────────────────────────────────


def _group_by_feature(test_cases: list[dict]) -> dict[str, list[dict]]:
    return _group_by_normalised_feature(test_cases)


def _safe_sheet_name(name: str, existing: set[str], max_len: int = 31) -> str:
    forbidden = r"\/*?:[]"
    clean = "".join(c if c not in forbidden else "_" for c in name)[:max_len]
    if not clean:
        clean = "Sheet"
    base, i = clean, 1
    while clean in existing:
        suffix = f"_{i}"
        clean = base[: max_len - len(suffix)] + suffix
        i += 1
    existing.add(clean)
    return clean


def _td_for_group(test_cases: list[dict], test_data_set: list[dict]) -> list[dict]:
    refs = {
        tc.get("test_data_ref", "").strip()
        for tc in test_cases
        if tc.get("test_data_ref")
    }
    return [td for td in test_data_set if td.get("id", "") in refs]


# ── Public API ────────────────────────────────────────────────────────────────


def to_excel(test_cases: list[dict], test_data_set: list[dict]) -> bytes:
    """
    Multi-sheet export — 1 sheet per functional module (feature_group normalised).

    Sheet layout:
      Sheet 1..N  : "<Feature name>"       — TC của từng chức năng
      Sheet N+1.. : "Data – <feature>"     — test data đi kèm (nếu có)
      Áp chót     : "📋 Quick Summary"     — bảng Sheet|Feature|TC|Ghi chú  ← MỚI
      Sheet cuối  : "📊 Coverage Summary"  — bảng tổng hợp chi tiết
    """
    from core.coverage_summary import (
        build_coverage_summary_sheet,
        build_quick_summary_sheet,
    )

    wb = Workbook()
    used_names: set[str] = set()

    # ── Sheet 1..N: Feature sheets ───────────────────────────────────────
    groups = _group_by_feature(test_cases)
    first_sheet = True

    # Lưu mapping feature_group → tên sheet thực tế (dùng cho Quick Summary)
    feature_to_sheet: dict[str, str] = {}

    for color_idx, (feature, tcs) in enumerate(groups.items()):
        color = _FEATURE_HDR_COLORS[color_idx % len(_FEATURE_HDR_COLORS)]
        sheet_title = _safe_sheet_name(feature[:31], used_names)
        feature_to_sheet[feature] = sheet_title

        if first_sheet:
            ws_feat = wb.active
            ws_feat.title = sheet_title
            first_sheet = False
        else:
            ws_feat = wb.create_sheet(sheet_title)

        _build_tc_sheet(ws_feat, tcs, header_color=color)

        td_subset = _td_for_group(tcs, test_data_set)
        if not td_subset and len(groups) == 1:
            td_subset = test_data_set

        if td_subset:
            td_sheet_title = _safe_sheet_name(f"Data – {feature[:20]}", used_names)
            ws_td_feat = wb.create_sheet(td_sheet_title)
            _build_td_sheet(ws_td_feat, td_subset)

    # ── Áp chót: Quick Summary ────────────────────────────────────────────
    qs_title = _safe_sheet_name("📋 Quick Summary", used_names)
    ws_quick = wb.create_sheet(qs_title)
    build_quick_summary_sheet(ws_quick, test_cases, sheet_name_map=feature_to_sheet)

    # ── Sheet cuối: Coverage Summary (chi tiết) ───────────────────────────
    summary_title = _safe_sheet_name("📊 Coverage Summary", used_names)
    ws_summary = wb.create_sheet(summary_title)
    build_coverage_summary_sheet(ws_summary, test_cases)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_excel_single(
    test_cases: list[dict],
    test_data_set: list[dict],
    sheet_name: str = "Test Cases",
) -> bytes:
    """Fallback 3-sheet export: TC + TD + Quick Summary + Coverage Summary."""
    from core.coverage_summary import (
        build_coverage_summary_sheet,
        build_quick_summary_sheet,
    )

    wb = Workbook()

    ws_tc = wb.active
    ws_tc.title = sheet_name
    _build_tc_sheet(ws_tc, test_cases)

    ws_td = wb.create_sheet("Test Data")
    _build_td_sheet(ws_td, test_data_set)

    ws_quick = wb.create_sheet("📋 Quick Summary")
    build_quick_summary_sheet(ws_quick, test_cases)

    ws_summary = wb.create_sheet("📊 Coverage Summary")
    build_coverage_summary_sheet(ws_summary, test_cases)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
