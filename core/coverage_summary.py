"""
core/coverage_summary.py
-------------------------
Sheet "📊 Coverage Summary" — bảng tổng hợp độ bao phủ, không biểu đồ.
Sheet "📋 Quick Summary"    — bảng ngắn gọn: Sheet | Feature | TC | Ghi chú
                               (đặt ở cuối, sau các sheet chức năng)
"""

from __future__ import annotations
from collections import defaultdict

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COVERAGE_META = {
    "P": {"label": "Positive", "color": "4CAF50"},
    "N": {"label": "Negative", "color": "F44336"},
    "B": {"label": "Boundary", "color": "2196F3"},
    "E": {"label": "Edge Case", "color": "FF9800"},
    "S": {"label": "Security", "color": "9C27B0"},
    "U": {"label": "UX", "color": "00BCD4"},
    "DB": {"label": "Database", "color": "795548"},
    "INT": {"label": "Integration", "color": "607D8B"},
}
COVERAGE_ORDER = ["P", "N", "B", "E", "S", "U", "DB", "INT"]

_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
_SUB_FILL = PatternFill("solid", fgColor="D6E4F0")
_BODY_FONT = Font(size=10, name="Arial")
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(vertical="center", horizontal="left")

_SCORE_FILLS = {
    (80, 101): PatternFill("solid", fgColor="C6EFCE"),
    (60, 80): PatternFill("solid", fgColor="FFEB9C"),
    (40, 60): PatternFill("solid", fgColor="FFCC7A"),
    (0, 40): PatternFill("solid", fgColor="FFC7CE"),
}


def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _score_fill(pct):
    for (lo, hi), fill in _SCORE_FILLS.items():
        if lo <= pct < hi:
            return fill
    return PatternFill("solid", fgColor="F0F0F0")


def _analyse(test_cases):
    overall = defaultdict(int)
    features = defaultdict(lambda: defaultdict(int))
    for tc in test_cases:
        ct = tc.get("coverage_type", "P").upper()
        if ct not in COVERAGE_META:
            ct = "P"
        fg = (tc.get("feature_group") or "General").strip()
        overall[ct] += 1
        features[fg][ct] += 1
        features[fg]["__total__"] += 1
        overall["__total__"] += 1
    return {
        "overall": dict(overall),
        "features": {k: dict(v) for k, v in features.items()},
    }


# ─────────────────────────────────────────────────────────────────────────────
# NEW: Quick Summary sheet  (Sheet | Feature | TC | Ghi chú)
# ─────────────────────────────────────────────────────────────────────────────

_QS_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_QS_HDR_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
_QS_TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
_QS_BODY_ALT = PatternFill("solid", fgColor="F5F9FF")  # alternate row

# Highlight labels for Ghi chú
_HIGHLIGHT_TYPES = {
    "S": "security",
    "B": "boundary",
    "E": "edge case",
    "DB": "DB",
    "INT": "integration",
    "U": "UX",
    "N": "negative",
    "P": "positive",
}
_CORE = {"P", "N", "B", "E", "S"}


def _build_ghi_chu(fd: dict, total: int) -> str:
    """
    Tạo chuỗi 'Ghi chú' từ dữ liệu 1 feature.
    Ví dụ: "Coverage 91% — security, boundary, RC timeout"
    """
    # Coverage score = % core types có ít nhất 1 TC
    covered = sum(1 for ct in _CORE if fd.get(ct, 0) > 0)
    score = round(covered / len(_CORE) * 100)

    # Danh sách highlights: các type có mặt (theo thứ tự ưu tiên)
    highlights = []
    for ct in ["P", "N", "S", "B", "E", "DB", "INT", "U"]:
        if fd.get(ct, 0) > 0:
            highlights.append(_HIGHLIGHT_TYPES[ct])

    if highlights:
        return f"Coverage {score}% — {', '.join(highlights)}"
    return f"Coverage {score}%"


def build_quick_summary_sheet(
    ws,
    test_cases: list[dict],
    sheet_name_map: dict[str, str] | None = None,
) -> None:
    """
    Xây dựng sheet "📋 Quick Summary" với layout:
        A: Sheet name  |  B: Feature (full name)  |  C: TC count  |  D: Ghi chú

    Parameters
    ----------
    ws            : openpyxl Worksheet đã được tạo sẵn.
    test_cases    : danh sách tất cả TC (sau khi sanitise).
    sheet_name_map: dict mapping feature_group → actual Excel sheet title.
                    Nếu None, dùng feature_group làm sheet name.
    """
    bdr = _border()

    # ── Tiêu đề lớn ──────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "📋 Quick Summary"
    title_cell.font = Font(bold=True, size=16, name="Arial", color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # ── Header row ───────────────────────────────────────────────────────
    HDR_ROW = 2
    headers = ["Sheet", "Feature", "TC", "Ghi chú"]
    widths = [22, 38, 8, 55]

    for ci, (hdr, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=HDR_ROW, column=ci, value=hdr)
        cell.fill = _QS_HDR_FILL
        cell.font = _QS_HDR_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[HDR_ROW].height = 24
    ws.freeze_panes = "A3"

    # ── Phân tích dữ liệu ─────────────────────────────────────────────
    data = _analyse(test_cases)
    features = data["features"]
    overall = data["overall"]
    feat_list = list(features.keys())  # giữ thứ tự xuất hiện trong TC

    # ── Data rows ─────────────────────────────────────────────────────
    DATA_START = HDR_ROW + 1

    for ri, feat in enumerate(feat_list):
        r = DATA_START + ri
        fd = features[feat]
        tot = fd.get("__total__", 0)

        # Sheet name: dùng map nếu có, fallback về feature name
        sheet_label = (sheet_name_map or {}).get(feat, feat)
        # Cắt nếu quá dài (Excel sheet name tối đa 31 ký tự)
        if len(sheet_label) > 31:
            sheet_label = sheet_label[:28] + "..."

        row_fill = _QS_BODY_ALT if ri % 2 == 1 else None  # alternate shading

        def _cell(col, val, align=None, bold=False, fill=None):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(bold=bold, size=10, name="Arial")
            c.alignment = align or Alignment(
                vertical="center", horizontal="left", wrap_text=True
            )
            c.border = bdr
            if fill:
                c.fill = fill
            elif row_fill:
                c.fill = row_fill
            return c

        _cell(
            1,
            sheet_label,
            align=Alignment(vertical="center", horizontal="left"),
            bold=True,
        )
        _cell(2, feat)
        _cell(3, tot, align=Alignment(vertical="center", horizontal="center"))
        _cell(4, _build_ghi_chu(fd, tot))

        ws.row_dimensions[r].height = 22

    # ── TOTAL row ────────────────────────────────────────────────────
    TOTAL_ROW = DATA_START + len(feat_list)
    grand = overall.get("__total__", 0)

    def _total_cell(col, val, align=None):
        c = ws.cell(row=TOTAL_ROW, column=col, value=val)
        c.font = Font(bold=True, size=11, name="Arial", color="1F4E79")
        c.fill = _QS_TOTAL_FILL
        c.alignment = align or Alignment(vertical="center", horizontal="left")
        c.border = bdr
        return c

    _total_cell(1, "TOTAL")
    _total_cell(2, f"{len(feat_list)} features")
    _total_cell(3, grand, align=Alignment(vertical="center", horizontal="center"))

    # Overall Ghi chú
    covered_all = sum(1 for ct in _CORE if overall.get(ct, 0) > 0)
    score_all = round(covered_all / len(_CORE) * 100)
    status_all = (
        "✅ Good"
        if score_all >= 80
        else "⚠️ Fair" if score_all >= 60 else "❌ Needs Work"
    )
    _total_cell(4, f"Overall Coverage {score_all}% — {status_all}")

    ws.row_dimensions[TOTAL_ROW].height = 26

    # Tab colour
    ws.sheet_properties.tabColor = "1F4E79"


# ─────────────────────────────────────────────────────────────────────────────
# Existing: Full Coverage Summary sheet (unchanged)
# ─────────────────────────────────────────────────────────────────────────────


def build_coverage_summary_sheet(ws, test_cases):
    data = _analyse(test_cases)
    features = data["features"]
    overall = data["overall"]
    bdr = _border()
    CORE = {"P", "N", "B", "E", "S"}

    # ── Tiêu đề ──────────────────────────────────────────────────────────
    last_col = get_column_letter(2 + len(COVERAGE_ORDER) + 2)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = "📊 Test Coverage Summary"
    c.font = Font(bold=True, size=16, name="Arial", color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"A2:{last_col}2")
    c2 = ws["A2"]
    c2.value = (
        f"Tổng test cases: {overall.get('__total__', 0)}   |   "
        f"Features/Modules: {len(features)}   |   "
        f"Coverage Score = % loại cốt lõi (P / N / B / E / S) có ít nhất 1 TC"
    )
    c2.font = Font(size=9, italic=True, name="Arial", color="555555")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Header bảng ──────────────────────────────────────────────────────
    col_labels = (
        ["Feature / Module", "Total TCs"]
        + [COVERAGE_META[ct]["label"] for ct in COVERAGE_ORDER]
        + ["Coverage\nScore", "Status"]
    )
    widths = [30, 10] + [12] * len(COVERAGE_ORDER) + [14, 14]

    HDR_ROW = 4
    for ci, (label, w) in enumerate(zip(col_labels, widths), 1):
        cell = ws.cell(row=HDR_ROW, column=ci, value=label)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[HDR_ROW].height = 30

    COLOR_ROW = 5
    ct_colors = (
        ["", ""] + [COVERAGE_META[ct]["color"] for ct in COVERAGE_ORDER] + ["", ""]
    )
    for ci, hex_c in enumerate(ct_colors, 1):
        cell = ws.cell(row=COLOR_ROW, column=ci, value="")
        if hex_c:
            cell.fill = PatternFill("solid", fgColor=hex_c)
        cell.border = bdr
    ws.row_dimensions[COLOR_ROW].height = 5

    ws.freeze_panes = f"A{COLOR_ROW + 1}"

    DATA_START = COLOR_ROW + 1
    feat_list = sorted(features.keys())

    for ri, feat in enumerate(feat_list):
        r = DATA_START + ri
        fd = features[feat]
        total = fd.get("__total__", 0)

        cell = ws.cell(row=r, column=1, value=feat)
        cell.font = Font(bold=True, size=10, name="Arial")
        cell.fill = _SUB_FILL
        cell.alignment = _LEFT
        cell.border = bdr

        cell = ws.cell(row=r, column=2, value=total)
        cell.font = _BODY_FONT
        cell.alignment = _CENTER
        cell.border = bdr

        for ci, ct in enumerate(COVERAGE_ORDER):
            count = fd.get(ct, 0)
            pct = round(count / total * 100, 1) if total > 0 else 0.0
            col = 3 + ci
            label = f"{count}\n({pct}%)" if count > 0 else "—"
            cell = ws.cell(row=r, column=col, value=label)
            cell.font = _BODY_FONT
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = bdr
            if count > 0:
                cell.fill = PatternFill(
                    "solid", fgColor=COVERAGE_META[ct]["color"] + "28"
                )

        covered = sum(1 for ct in CORE if fd.get(ct, 0) > 0)
        score = round(covered / len(CORE) * 100)
        SCORE_COL = 2 + len(COVERAGE_ORDER) + 1
        STATUS_COL = SCORE_COL + 1

        cell = ws.cell(row=r, column=SCORE_COL, value=f"{score}%")
        cell.font = Font(bold=True, size=11, name="Arial")
        cell.fill = _score_fill(score)
        cell.alignment = _CENTER
        cell.border = bdr

        status = (
            "✅ Good" if score >= 80 else "⚠️ Fair" if score >= 60 else "❌ Needs Work"
        )
        cell = ws.cell(row=r, column=STATUS_COL, value=status)
        cell.font = _BODY_FONT
        cell.alignment = _CENTER
        cell.border = bdr

        ws.row_dimensions[r].height = 30

    TOTAL_ROW = DATA_START + len(feat_list)
    grand = overall.get("__total__", 0)
    SCORE_COL = 2 + len(COVERAGE_ORDER) + 1
    STATUS_COL = SCORE_COL + 1
    ws.row_dimensions[TOTAL_ROW].height = 28

    cell = ws.cell(row=TOTAL_ROW, column=1, value="🔢 TOTAL")
    cell.font = Font(bold=True, size=11, name="Arial", color="1F4E79")
    cell.fill = _SUB_FILL
    cell.alignment = _LEFT
    cell.border = bdr

    cell = ws.cell(row=TOTAL_ROW, column=2, value=grand)
    cell.font = Font(bold=True, size=11, name="Arial")
    cell.fill = _SUB_FILL
    cell.alignment = _CENTER
    cell.border = bdr

    for ci, ct in enumerate(COVERAGE_ORDER):
        count = overall.get(ct, 0)
        pct = round(count / grand * 100, 1) if grand > 0 else 0.0
        label = f"{count}\n({pct}%)" if count > 0 else "—"
        cell = ws.cell(row=TOTAL_ROW, column=3 + ci, value=label)
        cell.font = Font(bold=True, size=10, name="Arial")
        cell.fill = _SUB_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = bdr

    covered_all = sum(1 for ct in CORE if overall.get(ct, 0) > 0)
    score_all = round(covered_all / len(CORE) * 100)

    cell = ws.cell(row=TOTAL_ROW, column=SCORE_COL, value=f"{score_all}%")
    cell.font = Font(bold=True, size=11, name="Arial")
    cell.fill = _score_fill(score_all)
    cell.alignment = _CENTER
    cell.border = bdr

    status_all = (
        "✅ Good"
        if score_all >= 80
        else "⚠️ Fair" if score_all >= 60 else "❌ Needs Work"
    )
    cell = ws.cell(row=TOTAL_ROW, column=STATUS_COL, value=status_all)
    cell.font = Font(bold=True, size=10, name="Arial")
    cell.fill = _SUB_FILL
    cell.alignment = _CENTER
    cell.border = bdr

    ws.sheet_properties.tabColor = "1F4E79"
